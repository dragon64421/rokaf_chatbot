from flask import Flask, request, jsonify
import openpyxl
import random
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)

def findxl(sheet, col, value): #엑셀파일의 특정 열에 값이 있는지 체크
    index = 1
    for row in sheet.rows:
        if str(value) in str(row[col].value) :
            return index
        index += 1
    return False

def findxl_ex(sheet, col, value):  # 엑셀파일의 특정 열에 값이 있는지 체크, 있으면 인덱스를 리스트로 반환, 완벽히 값이 같아야 함
    index = 1
    list = []
    for row in sheet.rows:
        if str(value) == str(row[col].value):
            list.append(index)
        index += 1
    return list

def signin(value, rcvmsg): #등록
    name_id = openpyxl.load_workbook("name_id.xlsx")
    name_id_sheet = name_id.worksheets[0]
    check_old = findxl(name_id_sheet, 5, value)
    if check_old:
        name_id.close()
        return 0
    else:
        if rcvmsg.startswith('등록'):
            if len(rcvmsg.split('.')) == 5:
                info = []
                for i in range(1, 5):
                    info.append(rcvmsg.split('.')[i])
                index = 1
                while name_id_sheet.cell(row=index, column = 1).value != None:
                    index += 1
                for i in range(1, 5):
                    name_id_sheet.cell(row=index, column=i, value = info[i-1])
                name_id_sheet.cell(row=index, column=5, value=0)
                name_id_sheet.cell(row=index, column=6, value=value)
                name_id_sheet.cell(row=index, column=7, value=0)
                name_id_sheet.cell(row=index, column=8, value=0)
                name_id.save('name_id.xlsx')
                name_id.close()
                return 1
            else:
                name_id.close()
                return 2
        else:
            name_id.close()
            return 2

def superadmin(mode, rcvmsg, kakaoid): #슈퍼관리자명령어
    name_id = openpyxl.load_workbook("name_id.xlsx")
    name_id_sheet = name_id.worksheets[0]
    check_old = findxl(name_id_sheet, 5, kakaoid)
    if name_id_sheet.cell(row=check_old, column=7).value == 2:
        if mode == 1: #관리자등록
            stnum = rcvmsg.split('.')[1]
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 2, stnum)
            name_id_sheet.cell(row=check_old, column=7, value=1)
            name_id.save('name_id.xlsx')
            return '등록완료'
        elif mode == 2: #관리자삭제
            stnum = rcvmsg.split('.')[1]
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 2, stnum)
            name_id_sheet.cell(row=check_old, column=7, value=0)
            name_id.save('name_id.xlsx')
            return '삭제완료'
        elif mode == 3: #마스터키
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            numvar = surveylist_sheet.cell(row=check_name, column=2).value
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            survey_file = openpyxl.load_workbook(surveyname + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            msg = ''
            index = 1
            while survey_sheet.cell(row=index, column=1).value != None:
                for i in range(1, 4+int(numvar)):
                    msg += str(survey_sheet.cell(row=index, column=i + 1).value)
                    msg += '.'
                msg += str(survey_sheet.cell(row=index, column=i + 2).value)
                msg += '\n'
                index += 1
            msg = msg[:-1]
            return msg
        elif mode == 4: #신청목록정보
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            index = 1
            while surveylist_sheet.cell(row=index, column=1).value != None:
                makerid = surveylist_sheet.cell(row=index, column=3).value
            msg = '미완성 명령어'
            return msg
    else:
        return '권한이 없습니다'

def admin(mode, rcvmsg, kakaoid): #관리자명령어
    name_id = openpyxl.load_workbook("name_id.xlsx")
    name_id_sheet = name_id.worksheets[0]
    check_old = findxl(name_id_sheet, 5, kakaoid)
    if name_id_sheet.cell(row=check_old, column=7).value >= 1:
        if mode == 999: #(구)신청생성, 작동안함(참고용코드)
            if len(rcvmsg.split('/')[0].split('.')) == 3 and len(rcvmsg.split('/')) >= 2 and len(rcvmsg.split('/')[1].split('.')) == int(rcvmsg.split('/')[0].split('.')[2]):
                newsruveyname = rcvmsg.split('.')[1]
                surveyform = rcvmsg.split('/')[1]
                if len(rcvmsg.split('/')) == 3:
                    surveyinfo = rcvmsg.split('/')[2]
                surveylist = openpyxl.load_workbook('surveylist.xlsx')
                surveylist_sheet = surveylist.worksheets[0]
                check_name = findxl(surveylist_sheet, 0, newsruveyname)
                if check_name == False:
                    index = 1
                    while surveylist_sheet.cell(row=index, column=1).value != None:
                        index += 1
                    surveylist_sheet.cell(row=index, column=1, value=newsruveyname)
                    surveylist_sheet.cell(row=index, column=2, value=rcvmsg.split('/')[0].split('.')[2])
                    surveylist_sheet.cell(row=index, column=3, value=kakaoid)
                    surveylist_sheet.cell(row=index, column=4, value=surveyform)
                    if len(rcvmsg.split('/')) == 3:
                        surveylist_sheet.cell(row=index, column=5, value=surveyinfo)
                    surveylist.save('surveylist.xlsx')
                    wb = openpyxl.Workbook()
                    wb.save(str(newsruveyname)+'.xlsx')
                    return '신청생성완료'
                else:
                    return '이미 존재하는 신청이름입니다. 다른 이름으로 생성해주세요'
            else:
                return '신청생성 양식 =\n 신청생성.신청이름.받을내용개수/양식/기타안내사항\nex) 신청생성.과일신청.2/과일종류.수량/과일종류: 사과, 수박, 복숭아\n***기타안내사항에 ./ 사용금지***'
        elif mode == 2: #신청삭제
            if len(rcvmsg.split('.')) == 2:
                deletesurveyname = rcvmsg.split('.')[1]
                surveylist = openpyxl.load_workbook('surveylist.xlsx')
                surveylist_sheet = surveylist.worksheets[0]
                check_name = findxl(surveylist_sheet, 0, deletesurveyname)
                if check_name == False:
                    return '존재하지 않는 신청이름입니다'
                elif kakaoid == surveylist_sheet.cell(row=check_name, column=3).value:
                    os.remove(deletesurveyname+'.xlsx')
                    surveylist_sheet.delete_rows(check_name)
                    surveylist.save('surveylist.xlsx')
                    return '삭제완료'
                else:
                    return '신청을 생성한 사람만 삭제할 수 있습니다'
            else:
                return '신청삭제 양식 =\n 신청삭제.신청이름'
        elif mode == 3: #내려받기
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            numvar = surveylist_sheet.cell(row=check_name, column=2).value
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name, column=7).value and kakaoid != surveylist_sheet.cell(row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name, column=9).value and kakaoid != surveylist_sheet.cell(row=check_name, column=10).value:
                return '해당 신청내용을 열람할 권한이 없습니다'
            survey_file = openpyxl.load_workbook(surveyname + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            msg = ''
            index = 1
            while survey_sheet.cell(row=index, column=1).value != None:
                for i in range(1, 4 + int(numvar)):
                    msg += str(survey_sheet.cell(row=index, column=i + 1).value)
                    msg += '.'
                msg += str(survey_sheet.cell(row=index, column=i + 2).value)
                msg += '\n'
                index += 1
            msg = msg[:-1]
            return msg
        elif mode == 4: #이메일받기
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name, column=7).value and kakaoid != surveylist_sheet.cell(row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name, column=9).value and kakaoid != surveylist_sheet.cell(row=check_name, column=10).value:
                return '해당 신청내용을 열람할 권한이 없습니다'
            email_user = 'afakakaobot@gmail.com'
            email_password = 'Qq28095774!'
            try:
                email_send = rcvmsg.split('.')[2] + '.' + rcvmsg.split('.')[3] + '.' + rcvmsg.split('.')[4]
            except:
                email_send = rcvmsg.split('.')[2] + '.' + rcvmsg.split('.')[3]
            print(email_send)
            subject = 'From. 행정이'
            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['To'] = email_send
            msg['Subject'] = subject
            body = '행정이가 보냄.'
            msg.attach(MIMEText(body, 'plain'))
            filename = surveyname + '.xlsx'
            attachment = open(filename, 'rb')
            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment", filename=os.path.basename(filename))
            msg.attach(part)
            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, email_password)
            server.sendmail(email_user, email_send, text)
            server.quit()
            return '이메일 전송완료'
        elif mode == 5:  # 교번순정렬
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            numvar = surveylist_sheet.cell(row=check_name, column=2).value
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name, column=7).value and kakaoid != surveylist_sheet.cell(row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name, column=9).value and kakaoid != surveylist_sheet.cell(row=check_name, column=10).value:
                return '해당 신청내용을 수정할 권한이 없습니다.'
            survey_file = openpyxl.load_workbook(surveyname + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            stid = []
            index = 1
            while survey_sheet.cell(row=index, column=1).value != None:
                stid.append(int(survey_sheet.cell(row=index, column=4).value))
                index += 1
            newfile = openpyxl.Workbook()
            newfilesheet = newfile.worksheets[0]
            for i in range(1, index):
                minindex = stid.index(min(stid))
                for a in range(1, 6+int(numvar)):
                    newfilesheet.cell(row=i, column=a).value = survey_sheet.cell(row=minindex+1, column=a).value
                stid[minindex] = 999999999999
            newfile.save(surveyname+'.xlsx')
            return '교번순 정렬완료'
        elif mode == 6:  # 교번없이내려받기
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            numvar = surveylist_sheet.cell(row=check_name, column=2).value
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name, column=7).value and kakaoid != surveylist_sheet.cell(row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name, column=9).value and kakaoid != surveylist_sheet.cell(row=check_name, column=10).value:
                return '해당 신청내용을 열람할 권한이 없습니다'
            survey_file = openpyxl.load_workbook(surveyname + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            msg = ''
            index = 1
            while survey_sheet.cell(row=index, column=1).value != None:
                msg += survey_sheet.cell(row=index, column=2).value
                msg += '.'
                msg += survey_sheet.cell(row=index, column=3).value
                msg += '.'
                msg += survey_sheet.cell(row=index, column=5).value
                msg += '.'
                for i in range(0, int(numvar)):
                    msg += str(survey_sheet.cell(row=index, column=6 + i).value)
                    msg += '.'
                msg = msg[:-1]
                msg += '\n'
                index += 1
            msg = msg[:-1]
            return msg
        elif mode == 7:  # 신청현황
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name, column=7).value and kakaoid != surveylist_sheet.cell(row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name, column=9).value and kakaoid != surveylist_sheet.cell(row=check_name, column=10).value:
                return '해당 신청내용을 열람할 권한이 없습니다'
            survey_file = openpyxl.load_workbook(surveyname + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            index = 1
            clscount = [0, 0, 0, 0]
            sqdcount = [0, 0, 0, 0, 0, 0, 0, 0]
            while survey_sheet.cell(row=index, column=1).value != None:
                if survey_sheet.cell(row=index, column=2).value == '69':
                    clscount[0] += 1
                if survey_sheet.cell(row=index, column=2).value == '70':
                    clscount[1] += 1
                if survey_sheet.cell(row=index, column=2).value == '71':
                    clscount[2] += 1
                if survey_sheet.cell(row=index, column=2).value == '72':
                    clscount[3] += 1
                if survey_sheet.cell(row=index, column=3).value == '1':
                    sqdcount[0] += 1
                if survey_sheet.cell(row=index, column=3).value == '2':
                    sqdcount[1] += 1
                if survey_sheet.cell(row=index, column=3).value == '3':
                    sqdcount[2] += 1
                if survey_sheet.cell(row=index, column=3).value == '4':
                    sqdcount[3] += 1
                if survey_sheet.cell(row=index, column=3).value == '5':
                    sqdcount[4] += 1
                if survey_sheet.cell(row=index, column=3).value == '6':
                    sqdcount[5] += 1
                if survey_sheet.cell(row=index, column=3).value == '7':
                    sqdcount[6] += 1
                if survey_sheet.cell(row=index, column=3).value == '8':
                    sqdcount[7] += 1
                index += 1
            msg = str(surveyname) + '의 신청수합현황'
            msg += '\n총 신청수: ' + str(index-1)
            for i in range(0, 4):
                if clscount[i] != 0:
                    msg += '\n' + str(69 + i) + '기: ' + str(clscount[i]) + '개'
            for i in range(0, 8):
                if sqdcount[i] != 0:
                    msg += '\n' + str(i + 1) + '중대: ' + str(sqdcount[i]) + '개'
            return msg
        elif mode == 8: # 권한부여
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            elif kakaoid == surveylist_sheet.cell(row=check_name, column=3).value:
                newsurveyadminclsnum = rcvmsg.split('.')[2]
                name_id = openpyxl.load_workbook("name_id.xlsx")
                name_id_sheet = name_id.worksheets[0]
                check_old = findxl(name_id_sheet, 2, newsurveyadminclsnum)
                newsurveyadminid = name_id_sheet.cell(row=check_old, column=6).value
                index = 1
                while surveylist_sheet.cell(row=check_name, column=5+index).value != None:
                    index += 1
                    if index == 6:
                        return '신청관리자는 최대 5명까지만 등록할 수 있습니다'
                surveylist_sheet.cell(row=check_name, column=5+index, value=newsurveyadminid)
                surveylist.save('surveylist.xlsx')
                return '신청관리자추가완료'
            else:
                return '신청을 생성한 사람만 신청관리자를 추가할 수 있습니다'
        elif mode == 9: # 권한삭제
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            elif kakaoid == surveylist_sheet.cell(row=check_name, column=3).value:
                delsurveyadminclsnum = rcvmsg.split('.')[2]
                name_id = openpyxl.load_workbook("name_id.xlsx")
                name_id_sheet = name_id.worksheets[0]
                check_old = findxl(name_id_sheet, 2, delsurveyadminclsnum)
                delsurveyadminid = str(name_id_sheet.cell(row=check_old, column=6).value)
                for i in range(0,5):
                    if str(surveylist_sheet.cell(row=check_name, column=6+i).value) == delsurveyadminid:
                        surveylist_sheet[check_name][5+i].value=None
                surveylist.save('surveylist.xlsx')
                return '신청관리자삭제완료'
            else:
                return '신청을 생성한 사람만 신청관리자를 삭제할 수 있습니다'
        elif mode == 10: #다운로드코드
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(
                    row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name,
                                                                                         column=7).value and kakaoid != surveylist_sheet.cell(
                    row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name,
                                                                                         column=9).value and kakaoid != surveylist_sheet.cell(
                    row=check_name, column=10).value:
                return '해당 신청내용을 열람할 권한이 없습니다'
            pw_list = openpyxl.load_workbook('pw_list.xlsx')
            pw_list_sheet = pw_list.worksheets[0]
            pw = random.randint(100000, 999999)
            pwcheck = 0
            while pwcheck != 0:
                index = 1
                while pw_list_sheet.cell(row=index, column=2).value != pw:
                    if pw_list_sheet.cell(row=index, column=1).value != None:
                        pwcheck = 1
                        break
                    else:
                        pw = random.randint(100000, 999999)
                        index += 1
            index = 1
            while pw_list_sheet.cell(row=index, column=1).value != None:
                index += 1
            filename = str(surveyname)+'.xlsx'
            pw_list_sheet.cell(row=index, column=1, value=filename)
            pw_list_sheet.cell(row=index, column=2, value=pw)
            pw_list.save('pw_list.xlsx')
            msg = '다운로드코드가 발급되었습니다.\nhttp://15.164.99.172에 접속하여\n'+str(pw)+'\n를 입력하면 다운로드가 시작됩니다.'
            return msg
        elif mode == 1: #신청생성
            if len(rcvmsg.split('/')[0].split('.')) == 3 and len(rcvmsg.split('/')) == 2 and len(
                    rcvmsg.split('/')[1].split('.')) == int(rcvmsg.split('/')[0].split('.')[2]):
                newsruveyname = rcvmsg.split('.')[1]
                surveyform = rcvmsg.split('/')[1]
                surveylist = openpyxl.load_workbook('surveylist.xlsx')
                surveylist_sheet = surveylist.worksheets[0]
                check_name = findxl(surveylist_sheet, 0, newsruveyname)
                if check_name == False:
                    index = 1
                    while surveylist_sheet.cell(row=index, column=1).value != None:
                        index += 1
                    surveylist_sheet.cell(row=index, column=1, value=newsruveyname)
                    surveylist_sheet.cell(row=index, column=2, value=rcvmsg.split('/')[0].split('.')[2])
                    surveylist_sheet.cell(row=index, column=3, value=kakaoid)
                    surveylist_sheet.cell(row=index, column=4, value=surveyform)
                    surveylist_sheet.cell(row=index, column=5, value=0)
                    surveylist.save('surveylist.xlsx')
                    wb = openpyxl.Workbook()
                    wb.save(str(newsruveyname) + '.xlsx')
                    return '신청생성완료'
                else:
                    return '이미 존재하는 신청이름입니다. 다른 이름으로 생성해주세요'
            else:
                return '신청생성 양식 =\n신청생성.신청이름.받을내용개수/양식\n선지형 질문: 질문[선지1,선지2,선지3]\nex) 신청생성.과일신청.2/과일종류[수박,사과,바나나].수량\n***질문구분= .  ***선지구분= ,'
        elif mode == 11: #신청마감
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(
                    row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name,
                                                                                         column=7).value and kakaoid != surveylist_sheet.cell(
                row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name,
                                                                                     column=9).value and kakaoid != surveylist_sheet.cell(
                row=check_name, column=10).value:
                return '해당 신청내용에 접근할 권한이 없습니다'
            survey = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_survey = findxl(surveylist_sheet, 0, survey)
            if surveylist_sheet.cell(row=check_survey, column=5).value == 1:
                return '이미 마감된 신청입니다'
            surveylist_sheet.cell(row=check_survey, column=5, value=1)
            surveylist.save('surveylist.xlsx')
            return '신청마감완료'
        elif mode == 12: #마감취소
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != surveylist_sheet.cell(row=check_name, column=3).value and kakaoid != surveylist_sheet.cell(
                    row=check_name, column=6).value and kakaoid != surveylist_sheet.cell(row=check_name,
                                                                                         column=7).value and kakaoid != surveylist_sheet.cell(
                row=check_name, column=8).value and kakaoid != surveylist_sheet.cell(row=check_name,
                                                                                     column=9).value and kakaoid != surveylist_sheet.cell(
                row=check_name, column=10).value:
                return '해당 신청내용에 접근할 권한이 없습니다'
            survey = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_survey = findxl(surveylist_sheet, 0, survey)
            if surveylist_sheet.cell(row=check_survey, column=5).value == 0:
                return '이미 열려있는 신청입니다'
            surveylist_sheet.cell(row=check_survey, column=5, value=0)
            surveylist.save('surveylist.xlsx')
            return '신청마감이 취소되었습니다'
        elif mode == 13:  # 투표생성
            if len(rcvmsg.split('/')[0].split('.')) == 3 and len(rcvmsg.split('/')) == 4 and len(
                    rcvmsg.split('/')[1].split('.')) == int(rcvmsg.split('/')[0].split('.')[2]):
                newvotename = rcvmsg.split('.')[1]
                voteform = rcvmsg.split('/')[1]
                votelist = openpyxl.load_workbook('votelist.xlsx')
                votelist_sheet = votelist.worksheets[0]
                check_name = findxl(votelist_sheet, 0, newvotename)
                if check_name == False:
                    index = 1
                    while votelist_sheet.cell(row=index, column=1).value != None:
                        index += 1
                    votelist_sheet.cell(row=index, column=1, value=newvotename)
                    votelist_sheet.cell(row=index, column=2, value=rcvmsg.split('/')[0].split('.')[2])
                    votelist_sheet.cell(row=index, column=3, value=kakaoid)
                    votelist_sheet.cell(row=index, column=4, value=voteform)
                    grade = rcvmsg.split('/')[2]
                    sqd = rcvmsg.split('/')[3]
                    name_id = openpyxl.load_workbook("name_id.xlsx")
                    name_id_sheet = name_id.worksheets[0]
                    wb = openpyxl.Workbook()
                    sheet = wb.worksheets[0]
                    index = 1
                    index2 = 1
                    while name_id_sheet.cell(row=index, column=1).value != None:
                        if grade == '전체':
                            if sqd == '전체':
                                sheet.cell(row=index2, column=1, value=name_id_sheet.cell(row=index, column=6).value)
                                index2 += 1
                            else:
                                if sqd == name_id_sheet.cell(row=index, column=2).value:
                                    sheet.cell(row=index2, column=1,
                                               value=name_id_sheet.cell(row=index, column=6).value)
                                    index2 += 1
                        else:
                            if sqd == '전체':
                                if grade == name_id_sheet.cell(row=index, column=1).value:
                                    sheet.cell(row=index2, column=1,
                                               value=name_id_sheet.cell(row=index, column=6).value)
                                    index2 += 1
                            else:
                                if grade == name_id_sheet.cell(row=index, column=1).value and sqd == name_id_sheet.cell(row=index, column=2).value:
                                    sheet.cell(row=index2, column=1,
                                               value=name_id_sheet.cell(row=index, column=6).value)
                                    index2 += 1
                        index += 1
                    sheet.cell(row=1, column=3, value='대상인원수')
                    sheet.cell(row=2, column=3, value='참여인원')
                    sheet.cell(row=3, column=3, value='미참여인원')
                    sheet.cell(row=1, column=4, value=index2-1)
                    sheet.cell(row=2, column=4, value=0)
                    sheet.cell(row=3, column=4, value=index2-1)
                    questionlist = (rcvmsg.split('/')[1]).split('.')
                    index2 = 5
                    for question in questionlist:
                        qs = question.split('[')[0]
                        sheet.cell(row=1, column=index2, value=qs)
                        selectlist = ((question.split('[')[1]).replace(']','')).split(',')
                        index = 2
                        for select in selectlist:
                            sheet.cell(row=index, column=index2, value=select)
                            sheet.cell(row=index, column=index2+1, value=0)
                            index += 1
                        index2 += 2
                    votelist.save('votelist.xlsx')
                    wb.save(str(newvotename) + '.xlsx')
                    return '투표생성완료'
                else:
                    return '이미 존재하는 투표이름입니다. 다른 이름으로 생성해주세요'
            else:
                return '투표생성 양식 =\n투표생성.투표이름.질문수/질문1[선지1,선지2,선지3].질문2[선지1,선지2,선지3]\n***질문구분= .  ***선지구분= ,'
        elif mode == 14: #투표삭제
            if len(rcvmsg.split('.')) == 2:
                deletevotename = rcvmsg.split('.')[1]
                votelist = openpyxl.load_workbook('votelist.xlsx')
                votelist_sheet = votelist.worksheets[0]
                check_name = findxl(votelist_sheet, 0, deletevotename)
                if check_name == False:
                    return '존재하지 않는 신청이름입니다'
                elif kakaoid == votelist_sheet.cell(row=check_name, column=3).value:
                    os.remove(deletevotename+'.xlsx')
                    votelist_sheet.delete_rows(check_name)
                    votelist.save('votelist.xlsx')
                    return '투표삭제완료'
                else:
                    return '투표를 생성한 사람만 삭제할 수 있습니다'
            else:
                return '투표삭제 양식 =\n 투표삭제.투표이름'
        elif mode == 15: #투표현황
            if len(rcvmsg.split('.')) != 2:
                return '투표현황 양식 =\n 투표현황,투표이름'
            votename = rcvmsg.split('.')[1]
            votelist = openpyxl.load_workbook('votelist.xlsx')
            votelist_sheet = votelist.worksheets[0]
            check_name = findxl(votelist_sheet, 0, votename)
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            if kakaoid != votelist_sheet.cell(row=check_name, column=3).value:
                return '투표를 생성한 사람만 내용을 확인할 수 있습니다.'
            msg = votename + '의 투표현황'
            vote_file = openpyxl.load_workbook(votename + '.xlsx')
            vote_sheet = vote_file.worksheets[0]
            voternum = str(vote_sheet.cell(row=1, column=4).value)
            voted = str(vote_sheet.cell(row=2, column=4).value)
            notvoted = str(vote_sheet.cell(row=3, column=4).value)
            msg += '\n\n대상인원수: ' + voternum + '\n참여인원수: ' + voted + '\n미참여인원수: ' + notvoted
            index2 = 6
            while vote_sheet.cell(row=1, column=index2-1).value != None:
                index = 2
                selectname = str(vote_sheet.cell(row=1, column=index2-1).value)
                msg += '\n\n' + selectname + '의 투표현황'
                while vote_sheet.cell(row=index, column=index2).value != None:
                    select = str(vote_sheet.cell(row=index, column=index2-1).value)
                    selectnum = str(vote_sheet.cell(row=index, column=index2).value)
                    msg += '\n' + select + ': ' + selectnum
                    index += 1
                index2 += 2
            return msg
    else:
        return '관리자권한이 없습니다\n관리자권한을 먼저 신청해주세요'

def normal(mode, rcvmsg, kakaoid):
    name_id = openpyxl.load_workbook("name_id.xlsx")
    name_id_sheet = name_id.worksheets[0]
    check_old = findxl(name_id_sheet, 5, kakaoid) #보낸사람 확인
    cls = name_id_sheet.cell(row=check_old, column=1).value
    sqd = name_id_sheet.cell(row=check_old, column=2).value
    clsnum = name_id_sheet.cell(row=check_old, column=3).value
    name = name_id_sheet.cell(row=check_old, column=4).value
    survey = rcvmsg.split('.')[1]
    surveylist = openpyxl.load_workbook('surveylist.xlsx')
    surveylist_sheet = surveylist.worksheets[0]
    check_survey = findxl(surveylist_sheet, 0, survey)
    if check_survey == False:
        return '없는 신청이름입니다'
    else:
        if mode == 1: #신청하기
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            if surveylist_sheet.cell(row=check_survey, column=5).value == 1:
                return '마감된 신청입니다.\n더이상의 신청은 불가합니다.'
            surveyname = surveylist_sheet.cell(row=check_survey, column=1).value
            numvar = surveylist_sheet.cell(row=check_survey, column=2).value
            surveyform = surveylist_sheet.cell(row=check_survey, column=4).value
            if len(rcvmsg.split('.')) == 2 + int(numvar):
                survey_file = openpyxl.load_workbook(survey+'.xlsx')
                survey_sheet = survey_file.worksheets[0]
                index = 1
                while survey_sheet.cell(row=index, column=1).value != None:
                    index += 1
                survey_sheet.cell(row=index, column=1, value=kakaoid)
                survey_sheet.cell(row=index, column=2, value=cls)
                survey_sheet.cell(row=index, column=3, value=sqd)
                survey_sheet.cell(row=index, column=4, value=clsnum)
                survey_sheet.cell(row=index, column=5, value=name)
                for i in range(1, int(numvar)+1):
                    survey_sheet.cell(row=index, column=5+i, value=rcvmsg.split('.')[1+i])
                survey_file.save(survey+'.xlsx')
                return '신청완료'
            else:
                return '해당 신청의 신청양식을 다시 확인해주세요\n[' + str(surveyname) + ']의 신청양식은\n[신청.' + str(surveyname) + '.' + str(
                    surveyform) + ']\n입니다'
        elif mode == 2: #삭제하기
            if surveylist_sheet.cell(row=check_survey, column=5).value == 1:
                return '마감된 신청입니다.\n신청내용을 수정할 수 없습니다.'
            survey_file = openpyxl.load_workbook(survey + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            check_info = findxl(survey_sheet, 0, kakaoid)
            while check_info != False:
                survey_sheet.delete_rows(check_info)
                survey_file.save(survey+'.xlsx')
                check_info = findxl(survey_sheet, 0, kakaoid)
            return '삭제완료'
        elif mode == 3: #내용확인
            surveyname = rcvmsg.split('.')[1]
            surveylist = openpyxl.load_workbook('surveylist.xlsx')
            surveylist_sheet = surveylist.worksheets[0]
            check_name = findxl(surveylist_sheet, 0, surveyname)
            numvar = surveylist_sheet.cell(row=check_name, column=2).value
            if check_name == False:
                return '존재하지 않는 신청이름입니다'
            survey_file = openpyxl.load_workbook(surveyname + '.xlsx')
            survey_sheet = survey_file.worksheets[0]
            msg = ''
            index = 1
            while survey_sheet.cell(row=index, column=1).value != None:
                if survey_sheet.cell(row=index, column=1).value == kakaoid:
                    for i in range(1, 4+int(numvar)):
                        msg += str(survey_sheet.cell(row=index, column=i + 1).value)
                        msg += '.'
                    msg += str(survey_sheet.cell(row=index, column=i + 2).value)
                    msg += '\n'
                index += 1
            msg = msg[:-1]
            return msg

def easysurvey(mode, rcvmsg, kakaoid):
    if mode == 0:  # 간편신청하기 모드 입장(신청목록 리스트로 출력, 신청자 상태 1로 변경)
        surveylist = openpyxl.load_workbook('surveylist.xlsx')
        surveylist_sheet = surveylist.worksheets[0]
        index = 1
        msglist = []
        while surveylist_sheet.cell(row=index, column=1).value != None:
            if surveylist_sheet.cell(row=index, column=5).value == 0:
                msglist.append([surveylist_sheet.cell(row=index, column=1).value, surveylist_sheet.cell(row=index, column=1).value])
            index += 1
        name_id = openpyxl.load_workbook("name_id.xlsx")
        name_id_sheet = name_id.worksheets[0]
        check_old = findxl(name_id_sheet, 5, kakaoid)
        name_id_sheet.cell(row=check_old, column=8, value=1)
        name_id.save('name_id.xlsx')
        return msglist
    if mode == 1:  # 신청을 선택함(기본정보 저장, 양식에서 첫번째 꺼 질문, 상태 신청이름.0 으로 변경)
        name_id = openpyxl.load_workbook("name_id.xlsx")
        name_id_sheet = name_id.worksheets[0]
        check_old = findxl(name_id_sheet, 5, kakaoid)  # 보낸사람 확인
        cls = name_id_sheet.cell(row=check_old, column=1).value
        sqd = name_id_sheet.cell(row=check_old, column=2).value
        clsnum = name_id_sheet.cell(row=check_old, column=3).value
        name = name_id_sheet.cell(row=check_old, column=4).value
        survey = rcvmsg
        surveylist = openpyxl.load_workbook('surveylist.xlsx')
        surveylist_sheet = surveylist.worksheets[0]
        check_survey = findxl(surveylist_sheet, 0, survey)
        if check_survey == False: # 신청이 중간에 삭제됐거나 마감된 경우 상태를 0으로 돌림
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=0)
            name_id.save('name_id.xlsx')
            return ['존재하지 않거나, 삭제된 신청입니다.']
        if surveylist_sheet.cell(row=check_survey, column=5).value == 1:
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=0)
            name_id.save('name_id.xlsx')
            return ['마감된 신청입니다']
        survey_file = openpyxl.load_workbook(survey + '.xlsx')
        survey_sheet = survey_file.worksheets[0]
        index = 1
        while survey_sheet.cell(row=index, column=1).value != None:
            index += 1
        survey_sheet.cell(row=index, column=1, value=kakaoid)
        survey_sheet.cell(row=index, column=2, value=cls)
        survey_sheet.cell(row=index, column=3, value=sqd)
        survey_sheet.cell(row=index, column=4, value=clsnum)
        survey_sheet.cell(row=index, column=5, value=name)
        survey_file.save(survey + '.xlsx')
        surveylist = openpyxl.load_workbook('surveylist.xlsx')
        surveylist_sheet = surveylist.worksheets[0]
        check_survey = findxl(surveylist_sheet, 0, survey)
        form = surveylist_sheet.cell(row=check_survey, column=4).value
        q1 = form.split('.')[0]
        if len(q1.split('[')) >= 2:
            msg = q1.split('[')[0]
            listdata = q1.split('[')[1]
            listdata = listdata.replace(']','')
            list = listdata.split(',')
            msgmsglist = [msg]
            for item in list:
                msgmsglist.append([item, item])
        else:
            msgmsglist = [q1]
        status = survey+'.0'
        name_id = openpyxl.load_workbook("name_id.xlsx")
        name_id_sheet = name_id.worksheets[0]
        check_old = findxl(name_id_sheet, 5, kakaoid)
        name_id_sheet.cell(row=check_old, column=8, value=status)
        name_id.save('name_id.xlsx')
        return msgmsglist
    else: #나머지 모든 작업 수행, status로 진행중인 신청과 진행도를 확인(신청이 삭제됐는지도 확인), 일단 입력값 저장, 방금입력한게 마지막 질문의 답인지 확인, 마지막이면 신청완료 아니면 다음 질문, 마지막이면 status 0 아니면 진행도 갱신
        survey = mode.split('.')[0]
        process = mode.split('.')[1]
        surveylist = openpyxl.load_workbook('surveylist.xlsx')
        surveylist_sheet = surveylist.worksheets[0]
        check_survey = findxl(surveylist_sheet, 0, survey)
        if check_survey == False:  # 신청이 중간에 삭제됐거나 (애러가 발생한 경우) 상태를 0으로 돌림
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=0)
            name_id.save('name_id.xlsx')
            return ['존재하지 않거나, 삭제된 신청입니다']
        if surveylist_sheet.cell(row=check_survey, column=5).value == 1:
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=0)
            name_id.save('name_id.xlsx')
            return ['마감된 신청입니다']
        survey_file = openpyxl.load_workbook(survey + '.xlsx')
        survey_sheet = survey_file.worksheets[0]
        index = 1
        while survey_sheet.cell(row=index, column=1).value != None:
            index += 1
        while index != 0:
            if survey_sheet.cell(row=index, column=1).value == kakaoid:
                survey_sheet.cell(row=index, column=6+int(process), value=rcvmsg)
                survey_file.save(survey + '.xlsx')
                break
            index += -1
        if index == 0:
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=0)
            name_id.save('name_id.xlsx')
            return ['애러발생(이전신청정보가 없음, 내용추가 불가능)\n처음부터 다시 진행해주세요. 지속적으로 같은 문제 발생시 문의바랍니다.']
        surveylist = openpyxl.load_workbook('surveylist.xlsx')
        surveylist_sheet = surveylist.worksheets[0]
        check_survey = findxl(surveylist_sheet, 0, survey)
        numvar = surveylist_sheet.cell(row=check_survey, column=2).value
        if int(process)+1 == int(numvar):
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=0)
            name_id.save('name_id.xlsx')
            return ['신청완료']
        process = int(process)
        process += 1
        status = survey + '.' + str(process)
        name_id = openpyxl.load_workbook("name_id.xlsx")
        name_id_sheet = name_id.worksheets[0]
        check_old = findxl(name_id_sheet, 5, kakaoid)
        name_id_sheet.cell(row=check_old, column=8, value=status)
        name_id.save('name_id.xlsx')
        survey_file.save(survey + '.xlsx')
        surveylist = openpyxl.load_workbook('surveylist.xlsx')
        surveylist_sheet = surveylist.worksheets[0]
        check_survey = findxl(surveylist_sheet, 0, survey)
        form = surveylist_sheet.cell(row=check_survey, column=4).value
        nextq = form.split('.')[process]
        if len(nextq.split('[')) >= 2:
            msg = nextq.split('[')[0]
            listdata = nextq.split('[')[1]
            listdata = listdata.replace(']','')
            list = listdata.split(',')
            msgmsglist = [msg]
            for item in list:
                msgmsglist.append([item, item])
        else:
            msgmsglist = [nextq]
        return msgmsglist

@app.route('/', methods=['POST'])
def message():
    dataReceive = request.get_json() #json으로 메시지 수신
    #print(dataReceive)
    rcvmsg = dataReceive["userRequest"]["utterance"] #수신 메시지 추출
    #print(rcvmsg)
    kakaoid = dataReceive["userRequest"]["user"]['id'] #카카오톡 id 추출
    signincheck = signin(kakaoid, rcvmsg)
    msglist = None
    if signincheck == 2:
        msg = '등록.기수.중대.교번.이름\n으로 먼저 등록을 해주세요'
    elif signincheck == 1:
        msg = '등록완료'
    elif signincheck == 0:
        name_id = openpyxl.load_workbook("name_id.xlsx")
        name_id_sheet = name_id.worksheets[0]
        check_old = findxl(name_id_sheet, 5, kakaoid)
        status = name_id_sheet.cell(row=check_old, column=8).value
        name_id.close()
        if status == 0 or status == None:
            if rcvmsg.startswith('관리자등록'):
                msg = superadmin(1, rcvmsg, kakaoid)
            elif rcvmsg.startswith('관리자삭제'):
                msg = superadmin(2, rcvmsg, kakaoid)
            elif rcvmsg.startswith('마스터키'):
                msg = superadmin(3, rcvmsg, kakaoid)
            elif rcvmsg.startswith('신청목록정보'):
                msg = superadmin(4, rcvmsg, kakaoid)
            elif rcvmsg.startswith('신청생성'):
                msg = admin(1, rcvmsg, kakaoid)
            elif rcvmsg.startswith('신청삭제'):
                msg = admin(2, rcvmsg, kakaoid)
            elif rcvmsg.startswith('내려받기'):
                msg = admin(3, rcvmsg, kakaoid)
            elif rcvmsg.startswith('이메일받기'):
                msg = admin(4, rcvmsg, kakaoid)
            elif rcvmsg.startswith('교번순정렬'):
                msg = admin(5, rcvmsg, kakaoid)
            elif rcvmsg.startswith('교번없이내려받기'):
                msg = admin(6, rcvmsg, kakaoid)
            elif rcvmsg.startswith('신청현황'):
                msg = admin(7, rcvmsg, kakaoid)
            elif rcvmsg.startswith('권한부여'):
                msg = admin(8, rcvmsg, kakaoid)
            elif rcvmsg.startswith('권한삭제'):
                msg = admin(9, rcvmsg, kakaoid)
            elif rcvmsg.startswith('다운로드코드'):
                msg = admin(10, rcvmsg, kakaoid)
            elif rcvmsg.startswith('신청마감'):
                msg = admin(11, rcvmsg, kakaoid)
            elif rcvmsg.startswith('마감취소'):
                msg = admin(12, rcvmsg, kakaoid)
            elif rcvmsg.startswith('신청'):
                msg = normal(1, rcvmsg, kakaoid)
            elif rcvmsg.startswith('삭제'):
                msg = normal(2, rcvmsg, kakaoid)
            elif rcvmsg.startswith('내용확인'):
                msg = normal(3, rcvmsg, kakaoid)
            elif rcvmsg.startswith('간편신청하기'): #간편신청하기 입력, 간편신청 목록 출력해야함
                msg = '👨현재 수합중인 신청목록입니다'
                msglist = easysurvey(0, rcvmsg, kakaoid)
            else:  # 폴백블록
                surveylist = openpyxl.load_workbook('surveylist.xlsx')
                surveylist_sheet = surveylist.worksheets[0]
                list = []
                ended = []
                a = 1
                while surveylist_sheet.cell(row=a, column=1).value != None:
                    if surveylist_sheet.cell(row=a, column=5).value == 0:
                        list.append(surveylist_sheet.cell(row=a, column=1).value)
                    elif surveylist_sheet.cell(row=a, column=5).value == 1:
                        ended.append(surveylist_sheet.cell(row=a, column=1).value)
                    a += 1
                msg = '👨행정업무 간소화를 위한 챗봇입니다.👨\n💻현재 수합중인 신청은\n' + str(
                    list) + '\n입니다.'
                if len(ended) != 0:
                    msg = msg + '\n👉' + str(ended) + '는 마감되었습니다'
                msg = msg + '\n👉프로필에서 자세한 사용법을 확인하실 수 있습니다'
                msglist =[['간편신청하기', '간편신청하기'],]
        elif rcvmsg == '나가기':
            status = 0
            name_id = openpyxl.load_workbook("name_id.xlsx")
            name_id_sheet = name_id.worksheets[0]
            check_old = findxl(name_id_sheet, 5, kakaoid)
            name_id_sheet.cell(row=check_old, column=8, value=status)
            name_id.save('name_id.xlsx')
            name_id.close()
            msg = '간편신청을 나갑니다\n지금까지 신청한 내역은 삭제되지 않습니다'
        elif status == 1: #간편신청목록에서 신청을 선택함
            msgmsglist = easysurvey(1, rcvmsg, kakaoid)
            msg = msgmsglist[0]
            if len(msgmsglist) >= 2:
                msglist = msgmsglist[1:]
        elif status == 2: #투표목록에서 투표를 선택함
            msgmsglist = []#vote(1, rcvmsg, kakaoid)
            msg = msgmsglist[0]
            if len(msgmsglist) >= 2:
                msglist = msgmsglist[1:]
        elif status.startswith('!투표'):
            msgmsglist = []#vote(status, rcvmsg, kakaoid)
            msg = msgmsglist[0]
            if len(msgmsglist) >= 2:
                msglist = msgmsglist[1:]
        else:
            msgmsglist = easysurvey(status, rcvmsg, kakaoid)
            msg = msgmsglist[0]
            if len(msgmsglist) >= 2:
                msglist = msgmsglist[1:]
    else:
        msg = 'unexpected error'

    if msglist == None or len(msglist)==0:
        #print(msg)
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": msg
                        }
                    }
                ]
            }}
    elif len(msglist) >= 1:
        #print(msg+str(msglist))
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": msg
                        }
                    }
                ],
                "quickReplies": []
            }}
        for item in msglist:
            dataSend["template"]["quickReplies"].append({
                "messageText": item[0],
                "action": "message",
                "label": item[1]
            })
    else:
        print('error(101)')
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": 'error(101)'
                        }
                    }
                ]
            }}

    return jsonify(dataSend)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=True)
