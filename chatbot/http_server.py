from flask import Flask, render_template, redirect, request, url_for, send_file
import openpyxl
import random
app = Flask(__name__)

@app.route('/')
def inputTest():
    return render_template('main.html')
    
@app.route('/check_pw',methods=['POST'])
def check_pw(pw =  None):
    pw = request.form['pw']
    pw_list = openpyxl.load_workbook("pw_list.xlsx")
    pw_list_sheet = pw_list.worksheets[0]
    index = 1
    verify = 0
    while pw_list_sheet.cell(row=index, column=1).value != None:
        if str(pw_list_sheet.cell(row=index, column=2).value) == str(pw):
            verify = 1
            break
        index += 1
    if request.method == 'POST' and verify == 1:
        filename = str(pw_list_sheet.cell(row=index, column=1).value)
        pw_list_sheet.delete_rows(index)
        pw_list.save('pw_list.xlsx')
        pw_list.close()
        return send_file(filename, attachment_filename=filename, as_attachment=True)
    else:
        return ('잘못된 인증코드입니다\n다시 입력하려면 뒤로가기 버튼을 눌러주세요')
    
 
if __name__ == '__main__':
    app.run(host="0.0.0.0", port=80, debug=True)
