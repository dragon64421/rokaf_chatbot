from flask import Flask, render_template, redirect, request, url_for, send_file
import openpyxl
app = Flask(__name__)

def findxl_ex(sheet, col, value):  # 엑셀파일의 특정 열에 값이 있는지 체크, 있으면 인덱스를 리스트로 반환, 완벽히 값이 같아야 함
    index = 1
    list = []
    for row in sheet.rows:
        if str(value) == str(row[col].value):
            list.append(index)
        index += 1
    return list

def add_file(filename, pw):
    pw_list = openpyxl.load_workbook("pw_list.xlsx")
    pw_list_sheet = pw_list.worksheets[0]
    if len(findxl_ex(pw_list_sheet,2,pw)) >= 1:
        print('이미 있음')
        return False
    last = pw_list_sheet.max_row
    pw_list_sheet[last + 1][1].value = filename
    pw_list_sheet[last + 1][2].value = pw
    pw_list.save('pw_list.xlsx')
    pw_list.close()
    return True

@app.route('/')
def inputTest():
    return render_template('main.html')
    
@app.route('/check_pw',methods=['POST'])
def check_pw(pw =  None):
    pw = request.form['pw']
    pw_list = openpyxl.load_workbook("pw_list.xlsx")
    pw_list_sheet = pw_list.worksheets[0]
    filename = findxl_ex(pw_list_sheet,2,pw)
    temp = filename[0]
    if request.method == 'POST' and len(filename) == 1:
        filename = pw_list_sheet[filename[0]][1].value
        pw_list_sheet.delete_rows(temp)
        pw_list.save('pw_list.xlsx')
        pw_list.close()
        return send_file(filename, attachment_filename=filename, as_attachment=True)
    else:
        return ('오류, 다시 입력')
    
 
if __name__ == '__main__':
    app.run(host="0.0.0.0", port=80, debug=True)
