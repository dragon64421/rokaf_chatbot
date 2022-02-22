from flask import Flask, request, jsonify
import openpyxl
import random
import os
import sendemail
from http_server import add_file

import plivo
try:
    client = plivo.RestClient(auth_id='',auth_token='')
except:
    print(1)

#엑셀 파일 load
name_xl = openpyxl.load_workbook("name_test.xlsx")
name_xl_sheet = name_xl.worksheets[0]
work_xl = openpyxl.load_workbook("work_test.xlsx")
work_xl_sheet = work_xl.worksheets[0]

app = Flask(__name__)

def intersection(a, b):
	c = []
	for item in a:
		if item in b:
			c.append(item)
	return c

def makeSubject(string):
    string = string.split(', ')
    returnString = ''
    for item in string:
        if '기' in item:
            lv = item.rstrip('기')
            for i in range(1,9):
                if lv+'.'+str(i) + ', ' in returnString:
                    continue
                else:
                    returnString += lv+'.'+str(i) + ', '
            continue
        if '중대' in item:
            sqd = item.rstrip('중대')
            for i in range(69,73):
                if str(i) + '.' + sqd + ', ' in returnString:
                    continue
                else:
                    returnString += str(i) + '.' + sqd + ', '
            continue
        else:
            returnString += item + ', '

    returnString = returnString.rstrip(', ')
    return returnString

def findxl(sheet, col, value): #엑셀파일의 특정 열에 값이 있는지 체크, 있으면 인덱스를 리스트로 반환
    index = 1
    list = []
    for row in sheet.rows:
        if str(value) in str(row[col].value) : list.append(index)
        index += 1
    return list

def findxl_ex(sheet, col, value):  # 엑셀파일의 특정 열에 값이 있는지 체크, 있으면 인덱스를 리스트로 반환, 완벽히 값이 같아야 함
    index = 1
    list = []
    for row in sheet.rows:
        if str(value) == str(row[col].value):
            list.append(index)
        index += 1
    return list

def send_sms(phone,code): #전화번호와 메시지로 문자 보내는 함수 => 이메일 보내는 함수로 수정 필요
    try:
        message_created = client.messages.create(
            src='+821034409041',
            dst=phone,
            text=code
        )
    except:
        return False
    return

def make_simple_reply(msg): #단순 답변을 만드는 함수
    dataSend = {
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "simpleText": {
                        "text": msg
                    }
                }
            ]
        }}
    print(msg)
    return dataSend

def make_quick_reply(msg,list): #quickReplies 답변을 만드는 함수
    dataSend = {
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "simpleText": {
                        "text": msg
                    }
                }
            ],
            "quickReplies": []
        }}
    for item in list:
        dataSend["template"]["quickReplies"].append({
            "messageText": item[0],
            "action": "message",
            "label": item[1]
        })
    print(msg+str(list))
    return dataSend

def get_work_list(id): #계급과 중대를 입력받아 할 수 있는 작업의 인덱스를 리턴하는 함수
    cadet_lv = id.split('.')[0] #id 에서 계급 추출
    cadet_sqd = id.split('.')[1] # id 에서 중대 추출
    work_index_list = findxl_ex(work_xl_sheet,2,'all') # 대상이 all 인 작업 검색
    work_index_list += findxl(work_xl_sheet, 2, str(cadet_lv)+'.'+str(cadet_sqd)) # 대상이 해당 기수.중대인 작업 검색
    work_index_list += findxl(work_xl_sheet, 2, name_xl_sheet[findxl_ex(name_xl_sheet,4,id)[0]][9].value) # 권한이 필요한 작업 검색
    work_index_list = list(set(work_index_list)) # 중복 인덱스 제거
    return work_index_list

def do_work(work_index,msg,id): # 특정 작업을 실행하는 함수
    work_index = int(work_index) # work_index type 변경
    name_index = findxl_ex(name_xl_sheet, 4, id)[0]
    name_xl_sheet[name_index][8].value = work_index + 1  # name_xl에서 상태 업데이트
    if msg == 'quit':
        name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
        return make_simple_reply('quit, 초기상태로 복귀') # 답장

    work_type = work_xl_sheet[work_index][3].value #작업의 종류

    if work_type == 'ask_str()':  # 해야 할 작업이 ask_date 이면
        return make_simple_reply(work_xl_sheet[work_index][4].value) # 메시지 답장

    # 해야 할 작업이 get 이면(형식 확인 절차 필요)
    if work_type == 'get_select()' or work_type == 'get_str()':
        filename = work_xl_sheet[work_index][6].value
        xl = openpyxl.load_workbook(filename+".xlsx") #작업에 필요한 xl 로드(예를들어 외출신청)
        xl_sheet = xl.worksheets[0]
        #order_list = findxl_ex(work_xl_sheet,6,work_xl_sheet[work_index][6].value) # 해당 작업이 전체 작업중에서 몇번째인지(외출신청 xl에서 몇번째 열에 데이터를 저장해야 하는지)
        #order = order_list.index(work_index) + 1
        order = 1 #해당 id의 마지막 열 검색
        xl_name_index = findxl_ex(xl_sheet,1,id)[0]
        while xl_sheet[xl_name_index][order].value != None:
            #print(order)
            order += 1
        xl_sheet[xl_name_index][order].value = str(msg) #외출신청 xl 에 데이터 저장
        xl.save(filename+".xlsx") #데이터 저장
        xl.close()
        return do_work(work_index+1,msg,id) #답장
    
    if work_type == 'cancel_work()': #신청 취소라면
        xl = openpyxl.load_workbook(msg+".xlsx") #작업에 필요한 xl 로드(예를들어 외출신청)
        xl_sheet = xl.worksheets[0]
        order = 2 # 1열이 이름이므로 신청내용은 2열부터 있음
        xl_name_index = findxl_ex(xl_sheet,1,id)[0] #새로 연 파일에서 이름 겁색
        while xl_sheet[xl_name_index][order].value != None: #신청 초기화
            xl_sheet[xl_name_index][order].value = None
            order += 1
        xl.save(msg+".xlsx") #저장
        xl.close()
        return do_work(work_index+1, msg, id)  # 답장

    if work_type == 'ask_select()':  # 해야 할 작업이 ask_select 이면
        ask_list = [] # make_quick_reply 에 전달할 리스트 생성
        ask_list_r = work_xl_sheet[work_index][5].value.split(',') # 엑셀파일에서 매개변수 가져오기
        for item in ask_list_r: # 리스트 생성
            ask_list.append([item,item])
        return make_quick_reply(work_xl_sheet[work_index][4].value,ask_list)
    
    if work_type == 'ask_select_work()':
        work_index_list = get_work_list(id) # 해당 생도가 할 수 있는 작업 목록
        work_list = [] # make_quick_reply 에 전달할 리스트 생성 
        for item in work_index_list: # 리스트 생성
            select = work_xl_sheet[item][1].value # 선택지
            if select == None:
                continue
            work_list.append([select, select])
        try: # 시스템 명령어 제거
            #work_list.pop(work_list.index(['관리', '관리']))
            work_list.pop(work_list.index(['신청 확인','신청 확인']))
            work_list.pop(work_list.index(['신청 취소', '신청 취소'])) #여기까지 일반사용자
            work_list.pop(work_list.index(['항목 추가', '항목 추가']))
            work_list.pop(work_list.index(['항목 제거', '항목 제거']))
            work_list.pop(work_list.index(['현황 파악', '현황 파악']))
            work_list.pop(work_list.index(['항목 초기화', '항목 초기화']))
            work_list.pop(work_list.index(['항목별 권한 관리', '항목별 권한 관리']))
            work_list.pop(work_list.index(['미실시자 파악', '미실시자 파악']))
            work_list.pop(work_list.index(['파일 다운로드', '파일 다운로드'])) #여기까지 일반관리자
            work_list.pop(work_list.index(['관리자 지정', '관리자 지정']))
            work_list.pop(work_list.index(['관리자 해제', '관리자 해제'])) # 여기까지 슈퍼관리자
        except:
            print('\n')
        return make_quick_reply('작업 선택',work_list)

    if work_type == 'get_select_work()':
        xl = openpyxl.load_workbook(str(msg)+".xlsx") #작업에 필요한 xl 로드(예를들어 외출신청)
        xl_sheet = xl.worksheets[0]
        #infoSize = xl_sheet.max_column # 최대 열 구하기(필요한 항목 개수)
        msg = ''
        i = 2
        xl_name_index = findxl_ex(xl_sheet,1,id)[0]
        while xl_sheet[xl_name_index][i].value != None:
            msg += str(xl_sheet[xl_name_index][i].value) + '/'
            i += 1
        msg = msg.rstrip('/')
        name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
        if i == 2 :
            return make_simple_reply('신청내역이 없습니다.')
        return make_simple_reply(msg)

    if work_type == 'add_work()': #중복 작업 방지 기능 추가 필요
        last = work_xl_sheet.max_row #work.xlsx 에서 마지막 행 구하기
        msg = msg.split('/')
        subject = makeSubject(msg[1]) # 대상자 추출
        if findxl_ex(work_xl_sheet,1,msg[0]): #이미 같은 이름의 작업이 있으면
             name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
             return make_simple_reply('이미 같은 이름의 신청이 있습니다. 다른 이름으로 다시 시도하세요.')
        work_xl_sheet[last + 1][1].value = msg[0] # 작업 이름
        work_xl_sheet[last + 1][2].value = subject # 대상자
        work_xl_sheet[last + 1][7].value = id # 추가자 id
        work_xl_sheet[last + 1][8].value = msg[2] # 추가 권한자
        new_xl = openpyxl.load_workbook("empty_work.xlsx") #새로운 엑셀파일 생성
        new_xl_sheet = new_xl.worksheets[0]
        for i in range(1,len(msg)-2): #work.xlsx 수정
            if msg[i + 2].startswith('#'): #질문이 선택형이면
                work_xl_sheet[last + (2 * i - 1)][3].value = 'ask_select()'
                msgs = msg[i+2].split('#')
                work_xl_sheet[last + (2 * i - 1)][4].value = msgs[1]
                work_xl_sheet[last + (2 * i - 1)][5].value = msgs[2]
            else:
                work_xl_sheet[last + (2 * i - 1)][3].value = 'ask_str()'
                work_xl_sheet[last + (2 * i - 1)][4].value = msg[i + 2]
            work_xl_sheet[last + (2 * i)][3].value = 'get_str()'
            work_xl_sheet[last + (2 * i)][6].value = msg[0]
        work_xl_sheet[last + 2*(len(msg)-3)+1][3].value = 'end'
        for i in range(3,len(msg)): # 새로운 엑셀파일 수정
            if msg[i].startswith('#'):
                new_xl_sheet[1][i - 1].value = msg[i].split('#')[1]
            else:
                new_xl_sheet[1][i-1].value = msg[i]
        index = 1
        if subject != 'all': # 대상자가 전체가 아니라면
            index = 1
            for row in new_xl_sheet.rows:  # 신규 엑셀 파일에서 대상자가 아니면 삭제
                if index < 2:
                    index += 1
                    continue
                if str(row[1].value)[:4] not in subject:
                    new_xl_sheet.delete_rows(index)
                    continue
                index += 1
        new_xl_sheet[1][18].value = None
        new_xl.save(msg[0]+'.xlsx')
        work_xl.save('work_test.xlsx')
        new_xl.close()
        return do_work(work_index+1, msg, id)  # 답장

    if work_type == 'delete_work()':
        index = findxl_ex(work_xl_sheet,1,msg) # 제거할 작업 검색
        if len(index) == 0: # 검색했는데 안나오면 
            name_xl_sheet[name_index][8].value = '0' #초기화
            return make_simple_reply('작업 없음, 다시시도')
        index = index[0]
        if work_xl_sheet[index][7].value == id: #권한 확인
            while work_xl_sheet[index][3].value != 'end': # work.xlsx 수정
                work_xl_sheet.delete_rows(index)
            work_xl_sheet.delete_rows(index)
            work_xl.save('work_test.xlsx')
            os.remove(msg+'.xlsx')
            return do_work(work_index+1, msg, id)  # 답장
        else: #권한 없으면
            name_xl_sheet[name_index][8].value = '0'  # 초기화
            return make_simple_reply('권한 없음, 다시시도')
    
    if work_type == 'add_manager()':
        msg = msg.split('/')
        if msg[1] not in '슈, 일':
            name_xl_sheet[name_index][8].value = '0'  # 초기화
            return make_simple_reply('잘못입력하셨습니다. 다시 입력해주세요.')
        index = findxl_ex(name_xl_sheet,4,msg[0])[0]
        name_xl_sheet[index][9].value = str(msg[1])
        return do_work(work_index+1, msg, id)  # 답장

    if work_type == 'delete_manager()':
        index = findxl_ex(name_xl_sheet, 4, msg)[0]
        name_xl_sheet[index][9].value = '평'
        return do_work(work_index+1, msg, id)  # 답장

    if work_type == 'file_download()':
        if str(id) in work_xl_sheet[findxl(work_xl_sheet,1,msg)[0]][7].value:
            pw = str(random.randint(100000,999999))
            if add_file(msg+'.xlsx', pw):
                name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
                return make_simple_reply('파일이 업로드 되었습니다.\n http://35.233.176.205 \n 위 주소로 접속하여 다운로드 코드 '+pw+'를 입력하여 파일을 다운로드하세요.')
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply('파일 추가 실패, 다시 시도해주세요.')
        name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
        return make_simple_reply('일치하는 파일 없음.')

    if work_type == 'get_list()':
        msg = msg.split('/')
        subject = msg[1]
        targetindex = findxl_ex(work_xl_sheet,1,msg[0])[0]
        check = False
        if name_xl_sheet[name_index][11].value != None:
            check = len(intersection(name_xl_sheet[name_index][11].value.split(','),str(work_xl_sheet[targetindex][8].value).split(','))) != 0
        if str(work_xl_sheet[targetindex][7].value) == id or id in str(work_xl_sheet[targetindex][8].value) or check: #권한 확인
            target_xl = openpyxl.load_workbook(msg[0]+".xlsx")
            target_xl_sheet = target_xl.worksheets[0]
            reply = ''
            for row in target_xl_sheet.rows:
                for col in row:
                    if col.value == None :
                        continue
                    if str(row[1].value)[:4] in subject or subject == 'all':
                        reply += str(col.value) + '/'
                reply = reply.rstrip('/')
                reply += '\n'
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply(reply)
        else:
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply('권한 없음')

    if work_type == 'get_list_not()':
        msg = msg.split('/')
        subject = msg[1]
        targetindex = findxl_ex(work_xl_sheet,1,msg[0])[0]
        check = False
        if name_xl_sheet[name_index][11].value != None:
            check = len(intersection(name_xl_sheet[name_index][11].value.split(','),str(work_xl_sheet[targetindex][8].value).split(','))) != 0
        if str(work_xl_sheet[targetindex][7].value) == id or id in str(work_xl_sheet[targetindex][8].value) or check: #권한 확인
            target_xl = openpyxl.load_workbook(msg[0]+".xlsx")
            target_xl_sheet = target_xl.worksheets[0]
            reply = ''
            for row in target_xl_sheet.rows:
                if row[2].value == None:
                    if str(row[1].value)[:4] in subject or subject == 'all':
                        reply += row[1].value
                        reply += '\n'
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply(reply)
        else:
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply('권한 없음')

    if work_type == 'permission_set()':
        msg = msg.split('/')
        subject = msg[1].split(',')
        targetindex = findxl_ex(work_xl_sheet,1,msg[0])[0]
        if str(work_xl_sheet[targetindex][7].value) == id: #권한 확인
            if msg[2] == '추가':
                for item in subject:
                    if item in str(work_xl_sheet[targetindex][8].value):
                        continue
                    else:
                        work_xl_sheet[targetindex][8].value = str(work_xl_sheet[targetindex][8].value) + ',' + item
            elif msg[2] == '제거':
                for item in subject:
                    if item in str(work_xl_sheet[targetindex][8].value):
                        work_xl_sheet[targetindex][8].value = str(work_xl_sheet[targetindex][8].value).replace(item, '')
            else:
                name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
                return make_simple_reply('오류, 추가 또는 제거라고 입력해주세요.')
        else:
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply('권한 없음')

        return do_work(work_index + 1, msg, id)  # 답장

    if work_type == 'reset_work()':
        targetindex = findxl_ex(work_xl_sheet,1,msg)[0]
        if str(work_xl_sheet[targetindex][7].value) == id: #권한 확인
            target_xl = openpyxl.load_workbook(msg+".xlsx")
            target_xl_sheet = target_xl.worksheets[0]
            index = 0
            for row in target_xl_sheet.rows:
                if index < 1:
                    index += 1
                    continue
                index1 = 0
                for col in row:
                    if index1 < 2:
                        index1 += 1
                        continue
                    col.value = None
            target_xl.save(msg+'.xlsx')
            return do_work(work_index+1, msg, id)  # 답장
        else:
            name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
            return make_simple_reply('권한 없음.')

    if work_type == 'end':
        name_xl_sheet[name_index][8].value = '0'  # name_xl에서 상태 업데이트
        try:
            return make_simple_reply(work_xl_sheet[work_index][4].value+'.')
        except:
            return make_simple_reply('완료!')

@app.route('/', methods=['POST'])
def message():
    dataReceive = request.get_json() #json으로 메시지 수신
    rcvmsg = dataReceive["userRequest"]["utterance"] #수신 메시지 추출

    kakaoid = dataReceive["userRequest"]["user"]['id'] #카카오톡 id 추출
    check_old = findxl_ex(name_xl_sheet, 6, kakaoid) #기존 회원인지 아닌지 체크

    if check_old: #기존 회원이면
        if name_xl_sheet[check_old[0]][7].value == '0': # 인증까지 끝났으면
            id = str(name_xl_sheet[check_old[0]][4].value)
            print(id + ',' + rcvmsg)
            if name_xl_sheet[check_old[0]][8].value == '0': #초기 상태이면
                work_index_list = get_work_list(name_xl_sheet[check_old[0]][4].value)  # 해당 생도가 할 수 있는 작업의 index get
                work_list = []
                for work in work_index_list:  # 작업 index로 list 생성
                    if work_xl_sheet[work][1].value == None:
                        continue
                    work_list.append([work_xl_sheet[work][1].value, work_xl_sheet[work][1].value])
                work_list.append(['관리', '관리'])
                try:  # 시스템 명령어 제거
                    work_list.pop(work_list.index(['신청 확인', '신청 확인']))
                    work_list.pop(work_list.index(['신청 취소', '신청 취소']))  # 여기까지 일반사용자
                    work_list.pop(work_list.index(['항목 추가', '항목 추가']))
                    work_list.pop(work_list.index(['항목 제거', '항목 제거']))
                    work_list.pop(work_list.index(['현황 파악', '현황 파악']))
                    work_list.pop(work_list.index(['항목 초기화', '항목 초기화']))
                    work_list.pop(work_list.index(['미실시자 파악', '미실시자 파악']))
                    work_list.pop(work_list.index(['항목별 권한 관리', '항목별 권한 관리']))
                    work_list.pop(work_list.index(['파일 다운로드', '파일 다운로드']))  # 여기까지 일반관리자
                    work_list.pop(work_list.index(['관리자 지정', '관리자 지정']))
                    work_list.pop(work_list.index(['관리자 해제', '관리자 해제']))  # 여기까지 슈퍼관리자
                except:
                    print('\n')
                name_xl_sheet[check_old[0]][8].value = '1'  # 작업 선택모드로 전환
                dataSend = make_quick_reply('작업선택', work_list)  # quick reply 생성


            elif name_xl_sheet[check_old[0]][8].value == '1': #작업 선택모드
                if rcvmsg == '관리':
                    work_index_list = get_work_list(name_xl_sheet[check_old[0]][4].value)  # 해당 생도가 할 수 있는 작업의 index get
                    work_list = []
                    set_list = [['신청 확인', '신청 확인'],['신청 취소', '신청 취소'],['항목 추가', '항목 추가'],['항목 제거', '항목 제거'],['현황 파악', '현황 파악'],
                                ['항목 초기화', '항목 초기화'],['미실시자 파악', '미실시자 파악'],['항목별 권한 관리', '항목별 권한 관리'],['파일 다운로드', '파일 다운로드'],['관리자 지정', '관리자 지정'],['관리자 해제', '관리자 해제']]
                    for work in work_index_list:  # 작업 index로 list 생성
                        if work_xl_sheet[work][1].value == None:
                            continue
                        work_list.append([work_xl_sheet[work][1].value, work_xl_sheet[work][1].value])
                    inter = intersection(work_list,set_list)
                    dataSend = make_quick_reply('작업선택', inter)  # quick reply 생성
                else:
                    work = findxl_ex(work_xl_sheet, 1, rcvmsg)  # 작업 검색
                    if len(work) == 0:  # 검색했는데 작업 없으면 (잘못 입력)
                        dataSend = make_simple_reply('잘못 입력하셨습니다. 다시 입력해주세요.')
                        name_xl_sheet[check_old[0]][8].value = '0'  # 초기화
                    else:
                        if work[0] in get_work_list(name_xl_sheet[check_old[0]][4].value):  # 권한 확인
                            name_xl_sheet[check_old[0]][8].value = work[0]  # 모드 전환
                            try:
                                dataSend = do_work(work[0], '', id)  # 작업 실행
                            except:
                                dataSend = make_simple_reply('알수없는 오류 발생. 처음부터 다시 해주세요.')
                                name_xl_sheet[check_old[0]][8].value = '0'  # 초기화
                        else:  # 권한 없으면
                            dataSend = make_simple_reply('권한 없음. 다시시도')
                            name_xl_sheet[check_old[0]][8].value = '0'  # 초기화

            else: #다른 작업 모드이면
                try:
                    dataSend = do_work(name_xl_sheet[check_old[0]][8].value, rcvmsg,id)
                except:
                    dataSend = make_simple_reply('알수없는 오류 발생. 처음부터 다시 해주세요.')
                    name_xl_sheet[check_old[0]][8].value = '0' 
            
        elif rcvmsg == name_xl_sheet[check_old[0]][7].value:  # 인증 안끝났으면(인증코드 일치하면)
            print(rcvmsg)
            name_xl_sheet[check_old[0]][7].value = '0' #인증이 완료되면 인증코드를 0으로 바꿈
            name_xl_sheet[check_old[0]][8].value = '0' #초기상태
            dataSend = make_simple_reply('인증완료. https://drive.google.com/file/d/1c1j1ZfMsJoC69wgJEKMU8duFpcLRhN9E/view?usp=sharing 에서 사용법을 확인하세요.')

        else : # 인증코드 불일치 -> 초기화
            print(rcvmsg)
            name_xl_sheet[check_old[0]][6].value = ''
            dataSend = make_simple_reply('인증 실패, 기수.중대.이름 다시 입력')

    else: #신규회원이면
        print(rcvmsg)
        name_of_new = findxl_ex(name_xl_sheet,4,rcvmsg) #메시지에서 이름 추출
        if name_of_new: # 명단에 이름이 있으면
            code = str(random.randint(1000, 9999)) #인증번호 생성
            name_xl_sheet[name_of_new[0]][7].value = code #인증번호 저장
            name_xl_sheet[name_of_new[0]][6].value = kakaoid #카카오 id 저장
            sendemail.send_email(name_xl_sheet[name_of_new[0]][5].value, '행정이 인증코드', '인증코드는' + code + '입니다.', None)
            dataSend = make_simple_reply('인증코드를 발송했습니다. 이메일 또는 문자를 확인하고 인증코드를 입력해주세요.')
            try:
                send_sms(name_xl_sheet[name_of_new[0]][10].value, '인증코드는' + code + '입니다.')  # 문자 발송
            except:
                print('\n')
        else: # 명단에 이름 없으면
            dataSend = make_simple_reply('기수.중대.이름이 틀렸습니다. 다시 입력하세요.')   
    
    name_xl.save("name_test.xlsx")
    work_xl.save('work_test.xlsx')
    return jsonify(dataSend)
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=True)
